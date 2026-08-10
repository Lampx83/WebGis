from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

FONT = "Arial"
COL = {
    "Người 1": "4472C4",
    "Người 2": "70AD47",
    "Người 3": "ED7D31",
    "Người 4": "7030A0",
}
LUONG = {
    "A": "A. Phần mềm WebGIS",
    "B": "B. Mua sắm / Đấu thầu (việc lớn)",
    "C": "C. Thực địa (khảo sát/lắp/chạy thử)",
    "D": "D. Viết báo cáo & tài liệu",
    "E": "E. Phân tích dữ liệu & Bài báo",
}
QUARTERS = [
    "Q4/2025", "Q1/2026", "Q2/2026", "Q3/2026", "Q4/2026",
    "Q1/2027", "Q2/2027", "Q3/2027", "Q4/2027",
    "Q1/2028", "Q2/2028", "Q3/2028", "Q4/2028",
]

# (STT, Mã, Luồng, Nội dung, Sản phẩm/Báo cáo, Phụ trách, Hỗ trợ, Bắt đầu, Deadline, qstart, qend)
TASKS = [
    # A. Phần mềm WebGIS
    (1, "8.1", "A", "Thiết kế WebGIS (giao diện + 4 nhóm tính năng)",
     "BC mô tả thiết kế WebGIS", "Người 1", "Người 4", "06/2026", "31/12/2026", 2, 4),
    (2, "8.2", "A", "Xây WebGIS tích hợp CSDL GIS + nền tảng IoT",
     "WebGIS tích hợp + BC mô tả", "Người 1", "Người 4", "06/2027", "31/12/2027", 6, 8),
    (3, "8.3", "A", "Thiết kế nền tảng dữ liệu IoT (data platform, MQTT, API trung gian)",
     "BC mô tả nền tảng IoT", "Người 1", "Người 2", "06/2027", "31/12/2027", 6, 8),
    (4, "9.2", "A", "Lập trình 3 kịch bản IoT + đặt ngưỡng + cảnh báo Zalo",
     "BC mô tả 3 kịch bản", "Người 3", "Người 2", "01/2028", "31/05/2028", 9, 10),
    (5, "9.3", "A", "Tích hợp dữ liệu trạm IoT thực vào WebGIS",
     "BC mô tả tích hợp CSDL", "Người 3", "Người 2", "03/2028", "31/08/2028", 9, 11),
    (6, "DEPLOY", "A", "Deploy & vận hành hệ thống (Docker/Portainer) — việc nhỏ",
     "Hệ thống chạy trên server", "Người 4", "Người 1", "11/2027", "31/03/2028", 8, 9),
    # B. Mua sắm / Đấu thầu (VIỆC LỚN)
    (7, "9.1", "B", "Phân tích, đánh giá & lựa chọn phương án lắp đặt IoT (SWOT)",
     "BC phương án IoT", "Người 2", "Người 3", "07/2026", "30/11/2026", 3, 4),
    (8, "B1", "B", "Tìm & liên hệ nhà cung cấp; khảo giá thị trường (≥3 báo giá)",
     "Bảng khảo giá + DS nhà cung cấp", "Người 2", "Người 1", "09/2026", "28/02/2027", 3, 5),
    (9, "B2", "B", "Lập danh mục thiết bị + yêu cầu kỹ thuật (specs) cho 3 kịch bản",
     "Danh mục + specs thiết bị", "Người 2", "Người 3", "11/2026", "31/01/2027", 4, 5),
    (10, "B3", "B", "Lập dự toán mua sắm thiết bị",
     "Dự toán mua sắm", "Người 2", "Thư ký Thảo", "01/2027", "31/03/2027", 5, 5),
    (11, "B4", "B", "Lập kế hoạch lựa chọn nhà thầu + hồ sơ mời thầu",
     "KH lựa chọn nhà thầu + HSMT", "Người 2", "Thư ký Thảo", "02/2027", "30/04/2027", 5, 6),
    (12, "B5", "B", "Tổ chức đấu thầu, đánh giá HSDT, ký hợp đồng",
     "QĐ trúng thầu + hợp đồng", "Người 2", "Thư ký Thảo; Người 1", "04/2027", "30/06/2027", 6, 6),
    (13, "B6", "B", "Nhận hàng, lắp đặt sơ bộ, nghiệm thu & thanh toán thiết bị",
     "Biên bản nghiệm thu thiết bị", "Người 2", "Người 3; TPI", "06/2027", "31/08/2027", 6, 7),
    # C. Thực địa
    (14, "2.3a", "C", "Khảo sát chọn vị trí trạm IoT (2 chuyến, 2 mùa; 3 TV×4 ngày)",
     "BC khảo sát vị trí lắp đặt", "Người 3", "Người 2, Người 4", "11/2025", "31/12/2026", 0, 4),
    (15, "2.3b", "C", "Lắp đặt trạm IoT tại hiện trường (1 chuyến; 5 TV×5 ngày)",
     "BC công tác lắp đặt trạm", "Người 3", "Người 2; Cảng Đình Vũ", "07/2027", "31/12/2027", 7, 8),
    (16, "2.3c", "C", "Kiểm tra máy + chạy thử thiết bị (1 chuyến; 5 TV×5 ngày)",
     "BC kiểm tra & chạy thử", "Người 3", "Người 4, Người 1", "01/2028", "30/09/2028", 9, 11),
    # E. Phân tích dữ liệu & Bài báo
    (17, "E1", "E", "Phân tích dữ liệu quan trắc IoT (thống kê; đánh giá ngưỡng & độ chính xác cảnh báo)",
     "BC phân tích dữ liệu IoT", "Người 2", "Người 3", "02/2028", "31/08/2028", 9, 11),
    (18, "E2", "E", "Phân tích & dự báo biến động không gian (đất, lớp phủ, khu nước, luồng cảng)",
     "BC phân tích biến động không gian", "Người 1", "Người 2", "01/2028", "31/08/2028", 9, 11),
    (19, "E3", "E", "Phân tích hiệu quả 3 kịch bản cảnh báo (đối chiếu số liệu thực địa)",
     "BC đánh giá kịch bản", "Người 3", "Người 2", "04/2028", "30/09/2028", 10, 11),
    (20, "PUB", "E", "01 bài báo: WebGIS tích hợp viễn thám + IoT quản lý cảng Hải Phòng",
     "01 bài báo (tạp chí/hội thảo)", "Người 1", "Người 2, Người 3", "06/2027", "31/10/2028", 8, 12),
    # D. Kiểm thử, tài liệu, chuyển giao
    (21, "9.4", "D", "Kết nối, vận hành thử nghiệm toàn hệ thống",
     "BC vận hành thử nghiệm", "Người 4", "Cả nhóm", "05/2028", "30/09/2028", 10, 11),
    (22, "9.5", "D", "Mô tả CSDL + biên soạn tài liệu hướng dẫn sử dụng",
     "Tài liệu HDSD", "Người 4", "Người 2", "06/2028", "30/09/2028", 10, 11),
    (23, "9.6", "D", "Tổ chức tập huấn dùng thử (tình nguyện viên)",
     "Biên bản tập huấn + BC kết quả", "Người 4", "Người 2", "05/2028", "31/10/2028", 10, 12),
    (24, "10.1", "D", "Chuyển giao, hướng dẫn vận hành cho 1 đơn vị",
     "Bộ tài liệu vận hành", "Người 4", "Người 1", "06/2028", "31/10/2028", 10, 12),
    (25, "10.2", "D", "Báo cáo đánh giá ứng dụng kết quả thử nghiệm",
     "BC đánh giá ứng dụng", "Người 4", "Đơn vị thử nghiệm", "08/2028", "31/10/2028", 11, 12),
    (26, "TH", "D", "Báo cáo tổng hợp phần Nội dung 8/9/10",
     "BC tổng hợp (nghiệm thu)", "Người 4", "Người 1 duyệt", "07/2028", "31/10/2028", 11, 12),
]

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F3864")
WRAP = Alignment(wrap_text=True, vertical="center")
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LUONG_FILL = {"A": "D9E1F2", "B": "FCE4D6", "C": "FFF2CC", "D": "E2EFDA", "E": "EDEDED"}

wb = Workbook()

# ── Sheet 1: Bảng công việc ──
ws = wb.active
ws.title = "Bảng công việc"
ws["A1"] = "BẢNG PHÂN CÔNG CÔNG VIỆC — NỘI DUNG 8/9/10 (WebGIS + IoT, cảng Hải Phòng)"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:J1")
ws["A2"] = "Chủ trì: Người 1 · Nhóm 4 người (gồm cả CN) · Giai đoạn 11/2025 – 10/2028"
ws["A2"].font = Font(name=FONT, italic=True, size=10, color="595959")
ws.merge_cells("A2:J2")

headers = ["STT", "Mã CV", "Luồng", "Nội dung công việc", "Sản phẩm / Báo cáo phải nộp",
           "Phụ trách (Lead)", "Hỗ trợ", "Bắt đầu", "Deadline", "Trạng thái"]
hr = 4
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=hr, column=c, value=h)
    cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CTR; cell.border = BORDER

r = hr + 1
for t in TASKS:
    stt, ma, lg, noidung, sp, lead, ho, bd, dl, qs, qe = t
    vals = [stt, ma, lg, noidung, sp, lead, ho, bd, dl, "Chưa bắt đầu"]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = BORDER
        cell.font = Font(name=FONT, size=10)
        cell.alignment = CTR if c in (1, 2, 3, 8, 9, 10) else WRAP
    if lead in COL:
        ws.cell(row=r, column=6).font = Font(name=FONT, size=10, bold=True, color=COL[lead])
    ws.cell(row=r, column=3).fill = PatternFill("solid", fgColor=LUONG_FILL[lg])
    ws.cell(row=r, column=9).font = Font(name=FONT, size=10, bold=True, color="C00000")
    r += 1

last = r - 1
dv = DataValidation(type="list", formula1='"Chưa bắt đầu,Đang làm,Hoàn thành,Trễ"', allow_blank=True)
ws.add_data_validation(dv); dv.add(f"J{hr+1}:J{last}")
widths = [5, 8, 8, 44, 34, 19, 22, 10, 11, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A5"; ws.row_dimensions[hr].height = 30

# ── Sheet 2: Ma trận phân công (người × công việc) ──
ms = wb.create_sheet("Ma trận phân công")
ms["A1"] = "MA TRẬN PHÂN CÔNG  (L = Chủ trì/Lead · H = Hỗ trợ)"
ms["A1"].font = TITLE_FONT
ms.merge_cells("A1:G1")
people_cols = ["Người 1", "Người 2", "Người 3", "Người 4"]
mh = ["Mã CV", "Luồng", "Nội dung công việc"] + people_cols
mhr = 3
for c, h in enumerate(mh, 1):
    cell = ms.cell(row=mhr, column=c, value=h)
    cell.fill = HDR_FILL
    cell.font = HDR_FONT if c <= 3 else Font(name=FONT, bold=True, color=COL[h], size=10)
    if c > 3:
        cell.fill = PatternFill("solid", fgColor="D6DCE4")
    cell.alignment = CTR; cell.border = BORDER

def role_of(person, lead, ho):
    if person == lead:
        return "L"
    # support match: person name appears in ho (handle "Người 1" alias)
    alias = "Người 1" if person == "Người 1" else person
    if person in ho or alias in ho or (person == "Người 1" and "Lâm" in ho):
        return "H"
    if "Cả nhóm" in ho:
        return "H"
    return ""

mr = mhr + 1
for t in TASKS:
    stt, ma, lg, noidung, sp, lead, ho, bd, dl, qs, qe = t
    ms.cell(row=mr, column=1, value=ma).alignment = CTR
    ms.cell(row=mr, column=1).font = Font(name=FONT, size=10, bold=True)
    lc = ms.cell(row=mr, column=2, value=lg); lc.alignment = CTR
    lc.fill = PatternFill("solid", fgColor=LUONG_FILL[lg])
    nc = ms.cell(row=mr, column=3, value=noidung); nc.alignment = WRAP
    nc.font = Font(name=FONT, size=9)
    for i, p in enumerate(people_cols):
        role = role_of(p, lead, ho)
        cell = ms.cell(row=mr, column=4 + i, value=role)
        cell.alignment = CTR; cell.border = BORDER
        if role == "L":
            cell.fill = PatternFill("solid", fgColor=COL[p])
            cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        elif role == "H":
            cell.fill = PatternFill("solid", fgColor="EDEDED")
            cell.font = Font(name=FONT, size=10, color="595959")
    for c in range(1, 4):
        ms.cell(row=mr, column=c).border = BORDER
    mr += 1

# count row
ms.cell(row=mr, column=3, value="Số CV chủ trì (L)").font = Font(name=FONT, bold=True, size=10)
ms.cell(row=mr, column=3).alignment = Alignment(horizontal="right")
_lc = {}
for _t in TASKS:
    _lc[_t[5]] = _lc.get(_t[5], 0) + 1
for i, p in enumerate(people_cols):
    cell = ms.cell(row=mr, column=4 + i, value=_lc.get(p, 0))
    cell.alignment = CTR; cell.font = Font(name=FONT, bold=True); cell.border = BORDER
ms.column_dimensions["A"].width = 8
ms.column_dimensions["B"].width = 7
ms.column_dimensions["C"].width = 50
for i in range(len(people_cols)):
    ms.column_dimensions[get_column_letter(4 + i)].width = 16
ms.freeze_panes = "D4"; ms.row_dimensions[mhr].height = 28

# ── Sheet 3: Timeline (Gantt) ──
gs = wb.create_sheet("Timeline (Gantt)")
gs["A1"] = "TIMELINE / GANTT — theo quý (tô màu theo người phụ trách)"
gs["A1"].font = TITLE_FONT
gs.merge_cells("A1:P1")
ghr = 3
gh = ["Mã CV", "Nội dung", "Lead"] + QUARTERS
for c, h in enumerate(gh, 1):
    cell = gs.cell(row=ghr, column=c, value=h)
    cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CTR; cell.border = BORDER
gr = ghr + 1
for t in TASKS:
    stt, ma, lg, noidung, sp, lead, ho, bd, dl, qs, qe = t
    gs.cell(row=gr, column=1, value=ma).alignment = CTR
    gs.cell(row=gr, column=1).font = Font(name=FONT, size=10, bold=True)
    nc = gs.cell(row=gr, column=2, value=noidung); nc.alignment = WRAP; nc.font = Font(name=FONT, size=9)
    lc = gs.cell(row=gr, column=3, value=lead); lc.alignment = CTR
    lc.font = Font(name=FONT, size=9, bold=True, color=COL.get(lead, "000000"))
    for c in range(1, 4):
        gs.cell(row=gr, column=c).border = BORDER
    color = COL.get(lead, "808080")
    for qi in range(len(QUARTERS)):
        cell = gs.cell(row=gr, column=4 + qi); cell.border = BORDER
        if qs <= qi <= qe:
            cell.fill = PatternFill("solid", fgColor=color)
    gr += 1
gs.column_dimensions["A"].width = 8
gs.column_dimensions["B"].width = 42
gs.column_dimensions["C"].width = 18
for qi in range(len(QUARTERS)):
    gs.column_dimensions[get_column_letter(4 + qi)].width = 8
gs.freeze_panes = "D4"; gs.row_dimensions[ghr].height = 28
lgr = gr + 2
gs.cell(row=lgr, column=1, value="Chú thích:").font = Font(name=FONT, bold=True, size=10)
for i, (name, color) in enumerate(COL.items()):
    cell = gs.cell(row=lgr + 1 + i, column=1)
    cell.fill = PatternFill("solid", fgColor=color); cell.border = BORDER
    gs.cell(row=lgr + 1 + i, column=2, value=name).font = Font(name=FONT, size=10)

# ── Sheet 4: Tổng hợp theo người ──
ss = wb.create_sheet("Tổng hợp theo người")
ss["A1"] = "TỔNG HỢP KHỐI LƯỢNG THEO NGƯỜI"
ss["A1"].font = TITLE_FONT
ss.merge_cells("A1:D1")
sh = ["Người", "Số CV chủ trì (Lead)", "Vai trò / Luồng chính", "Sản phẩm / Báo cáo chính"]
for c, h in enumerate(sh, 1):
    cell = ss.cell(row=3, column=c, value=h)
    cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CTR; cell.border = BORDER
people = [
    ("Người 1", "Chủ trì + Phần mềm WebGIS (A) + Phân tích biến động KG + Bài báo (E)",
     "BC 8.1/8.2/8.3 · BC biến động không gian · 01 bài báo"),
    ("Người 2", "Mua sắm/Đấu thầu (B, việc lớn) + Phân tích dữ liệu IoT (E)",
     "BC 9.1 · Dự toán/HSMT/HĐ · Nghiệm thu TB · BC phân tích IoT"),
    ("Người 3", "Thực địa (C) + Kịch bản IoT (A) + Phân tích kịch bản (E)",
     "BC khảo sát/lắp đặt · BC 9.2/9.3 · BC đánh giá kịch bản"),
    ("Người 4", "Kiểm thử, vận hành, tài liệu, chuyển giao (D) + Deploy",
     "BC 9.4/9.6/10.2 · Tài liệu HDSD · Deploy · BC tổng hợp"),
]
lead_count = {}
for t in TASKS:
    lead_count[t[5]] = lead_count.get(t[5], 0) + 1
sr = 4
for name, role, rep in people:
    ss.cell(row=sr, column=1, value=name).font = Font(name=FONT, size=10, bold=True, color=COL[name])
    cc = ss.cell(row=sr, column=2, value=lead_count.get(name, 0)); cc.alignment = CTR
    cc.font = Font(name=FONT, size=10, bold=True)
    ss.cell(row=sr, column=3, value=role).alignment = WRAP
    ss.cell(row=sr, column=4, value=rep).alignment = WRAP
    for c in range(1, 5):
        ss.cell(row=sr, column=c).border = BORDER
        if c in (3, 4):
            ss.cell(row=sr, column=c).font = Font(name=FONT, size=10)
    sr += 1
ss.cell(row=sr, column=1, value="TỔNG").font = Font(name=FONT, bold=True)
ss.cell(row=sr, column=2, value=sum(lead_count.values())).alignment = CTR
ss.cell(row=sr, column=2).font = Font(name=FONT, bold=True)
for c in range(1, 5):
    ss.cell(row=sr, column=c).border = BORDER
ss.cell(row=sr + 2, column=1,
        value="Ghi chú: Mua sắm/đấu thầu (B) là khối việc lớn, tập trung 2026 H2 – 2027 H2, do Người 2 chủ trì, "
              "phối hợp Thư ký Thảo (dự toán/đấu thầu) và TPI (tư vấn kỹ thuật).").font = Font(name=FONT, italic=True, size=9, color="595959")
ss.merge_cells(start_row=sr + 2, start_column=1, end_row=sr + 2, end_column=4)
ss.column_dimensions["A"].width = 20
ss.column_dimensions["B"].width = 20
ss.column_dimensions["C"].width = 52
ss.column_dimensions["D"].width = 52

out = "/Users/mac/Documents/Code/WebGis/KeHoach_PhanCong_ND8-9-10.xlsx"
wb.save(out)
print("Saved:", out)
print("Lead counts:", lead_count, "Total:", sum(lead_count.values()))
